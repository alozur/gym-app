"""
Generate SQL import file for gym-tracker historical data.
Reads all 5 Excel routine files and produces:
  1. INSERT statements for exercises (ON CONFLICT skip)
  2. INSERT statements for workout_templates + template_exercises
  3. INSERT statements for workout_sessions + workout_sets + exercise_progress
"""
import openpyxl
import re
import uuid
from datetime import datetime, timedelta
from decimal import Decimal

SCHEMA = "gym"
USER_EMAIL = "alonsozurera@gmail.com"

# ── Excel file config ───────────────────────────────────────────────────────
FILES = [
    {
        "path": r"C:\Users\ZureraA\Downloads\1 Upper Day.xlsx",
        "guide_sheet": "Guide",
        "template_name": "1 Upper Day",
        "title_row_col": (1, 2),  # C1
    },
    {
        "path": r"C:\Users\ZureraA\Downloads\2 Lower Day (1).xlsx",
        "guide_sheet": "Guide",
        "template_name": "2 Lower Day",
        "title_row_col": (1, 2),
    },
    {
        "path": r"C:\Users\ZureraA\Downloads\3 Pull Day.xlsx",
        "guide_sheet": "Pull",
        "template_name": "3 Pull Day",
        "title_row_col": (1, 2),
    },
    {
        "path": r"C:\Users\ZureraA\Downloads\4 Push Day.xlsx",
        "guide_sheet": "Guide",
        "template_name": "4 Push Day",
        "title_row_col": (1, 2),
    },
    {
        "path": r"C:\Users\ZureraA\Downloads\5 Legs Day.xlsx",
        "guide_sheet": "Guide",
        "template_name": "5 Legs Day",
        "title_row_col": (1, 2),
    },
]

# ── Exercise → muscle_group + equipment mapping ────────────────────────────
EXERCISE_INFO = {
    # Upper Day
    "45° Incline Barbell Press": ("Chest", "Barbell"),
    "Cable Crossover Ladder": ("Chest", "Cable"),
    "Wide-Grip Pull-Up": ("Back", "Bodyweight"),
    "High-Cable Lateral Raise": ("Shoulders", "Cable"),
    "Pendlay Deficit Row": ("Back", "Barbell"),
    "Overhead Cable Triceps Extension (Bar)": ("Triceps", "Cable"),
    "Bayesian Cable Curl": ("Biceps", "Cable"),
    "45° Incline DB Press": ("Chest", "Dumbbell"),
    # Lower Day
    "Lying Leg Curl": ("Hamstrings", "Machine"),
    "Smith Machine Squat": ("Quads", "Smith Machine"),
    "Barbell RDL": ("Hamstrings", "Barbell"),
    "Leg Extension": ("Quads", "Machine"),
    "Standing Calf Raise": ("Calves", "Machine"),
    "Cable Crunch": ("Abs", "Cable"),
    "Dumbell RDL": ("Hamstrings", "Dumbbell"),
    "Leg Press Calf Press": ("Calves", "Machine"),
    # Pull Day
    "Neutral-Grip Lat Pulldown": ("Back", "Cable"),
    "Chest-Supported Machine Row": ("Back", "Machine"),
    "Neutral-Grip Seated Cable Row": ("Back", "Cable"),
    "1-Arm 45° Cable Rear Delt Flye": ("Shoulders", "Cable"),
    "Machine Shrug": ("Traps", "Machine"),
    "EZ-Bar Cable Curl": ("Biceps", "Cable"),
    "Machine Preacher Curl": ("Biceps", "Machine"),
    "DB Shrug": ("Traps", "Dumbbell"),
    # Push Day
    "Barbell Bench Press": ("Chest", "Barbell"),
    "Machine Shoulder Press": ("Shoulders", "Machine"),
    "Bottom-Half DB Flye": ("Chest", "Dumbbell"),
    "Overhead CableTriceps Extension (Bar)": ("Triceps", "Cable"),
    "Cable Triceps Kickback": ("Triceps", "Cable"),
    "Roman Chair Leg Raise": ("Abs", "Bodyweight"),
    "Pec Deck": ("Chest", "Machine"),
    # Legs Day
    "Leg Press": ("Quads", "Machine"),
    "Seated Leg Curl": ("Hamstrings", "Machine"),
    "DB Bulgarian Split Squat": ("Quads", "Dumbbell"),
    "Machine Hip Adduction": ("Adductors", "Machine"),
    "Machine Hip Abduction": ("Abductors", "Machine"),
    # ── Substitution exercises ──
    "45° Incline Machine Press": ("Chest", "Machine"),
    "Wide-Grip Lat Pulldown": ("Back", "Cable"),
    "Dual-Handle Lat Pulldown": ("Back", "Cable"),
    "High-Cable Cuffed Lateral Raise": ("Shoulders", "Cable"),
    "Lean-In DB Lateral Raise": ("Shoulders", "Dumbbell"),
    "Smith Machine Row": ("Back", "Smith Machine"),
    "Single-Arm DB Row": ("Back", "Dumbbell"),
    "Overhead Cable Triceps Extension (Rope)": ("Triceps", "Cable"),
    "DB Skull Crusher": ("Triceps", "Dumbbell"),
    "Seated Super-Bayesian High Cable Curl": ("Biceps", "Cable"),
    "Incline DB Stretch Curl": ("Biceps", "Dumbbell"),
    "Neutral-Grip Pull-Up": ("Back", "Bodyweight"),
    "Chest-Supported T-Bar Row": ("Back", "Barbell"),
    "Incline Chest-Supported DB Row": ("Back", "Dumbbell"),
    "Helms Row": ("Back", "Dumbbell"),
    "Rope Face Pull": ("Shoulders", "Cable"),
    "Cable Paused Shrug-In": ("Traps", "Cable"),
    "EZ-Bar Curl": ("Biceps", "Barbell"),
    "EZ-Bar Preacher Curl": ("Biceps", "Barbell"),
    "Machine Chest Press": ("Chest", "Machine"),
    "DB Bench Press": ("Chest", "Dumbbell"),
    "Cable Shoulder Press": ("Shoulders", "Cable"),
    "Seated DB Shoulder Press": ("Shoulders", "Dumbbell"),
    "Bottom-Half Seated Cable Flye": ("Chest", "Cable"),
    "Low-to-High Cable Crossover": ("Chest", "Cable"),
    "DB TricepsKickback": ("Triceps", "Dumbbell"),
    "Bench Dip": ("Triceps", "Bodyweight"),
    "Hanging Leg Raise": ("Abs", "Bodyweight"),
    "Modified Candlestick": ("Abs", "Bodyweight"),
    "Smith Machine Static Lunge": ("Quads", "Smith Machine"),
    "DB Walking Lunge": ("Quads", "Dumbbell"),
    "Nordic Ham Curl": ("Hamstrings", "Bodyweight"),
    "DB Step-Up": ("Quads", "Dumbbell"),
    "Goblet Squat": ("Quads", "Dumbbell"),
    "Reverse Nordic": ("Quads", "Bodyweight"),
    "Sissy Squat": ("Quads", "Bodyweight"),
    "Cable Hip Adduction": ("Adductors", "Cable"),
    "Copenhagen Hip Adduction": ("Adductors", "Bodyweight"),
    "Cable Hip Abduction": ("Abductors", "Cable"),
    "Lateral Band Walk": ("Abductors", "Band"),
    "Seated Calf Raise": ("Calves", "Machine"),
    "DB Curl": ("Biceps", "Dumbbell"),
    "DB Preacher Curl": ("Biceps", "Dumbbell"),
    "DB RDL": ("Hamstrings", "Dumbbell"),
    "Decline Weighted Crunch": ("Abs", "Dumbbell"),
    "High-Bar Back Squat": ("Quads", "Barbell"),
    "Machine Crunch": ("Abs", "Machine"),
    "Meadows Row": ("Back", "Barbell"),
    "Reverse Pec Deck": ("Shoulders", "Machine"),
    "Snatch-Grip RDL": ("Hamstrings", "Barbell"),
}

# ── Name normalization for matching sheet names to exercise names ──────────
# Some sheet names are truncated or slightly different from the Guide
SHEET_NAME_MAP = {
    "Overhead Cable Triceps Extensio": "Overhead Cable Triceps Extension (Bar)",
    "Overhead CableTriceps Extension": "Overhead CableTriceps Extension (Bar)",
}

SKIP_SHEETS = {
    "Stats", "WarmUpNotes", "Guide", "Pull", "Template Exercises",
    "Copia de LowerW1",
}


def sql_str(v):
    """Escape a string for SQL."""
    if v is None:
        return "NULL"
    s = str(v).replace("'", "''")
    return f"'{s}'"


def sql_num(v):
    if v is None:
        return "NULL"
    return str(v)


def parse_rpe(s):
    """Parse RPE string like '8-9' into (min, max) or ('N/A','N/A')."""
    if not s or s == "N/A":
        return (None, None)
    s = str(s).strip()
    if "-" in s:
        parts = s.split("-")
        return (parts[0].strip(), parts[1].strip())
    return (s, s)


def is_dash_value(v):
    """Check if a value is a dash-surrounded number (different machine, skip)."""
    if v is None:
        return False
    s = str(v).strip()
    # Matches patterns like '-13,6-', '--30-', '-50-'
    if re.match(r"^-[\d,.\-]+-$", s):
        return True
    # Also check for patterns like '--30-' or '--40-'
    if re.match(r"^--\d+[\d,.]*-$", s):
        return True
    # Strings with asterisks like '-22,5*'
    if "*" in s and s.startswith("-"):
        return False  # This is a real negative weight with note
    return False


def is_valid_weight(v):
    """Check if a value is a valid numeric weight (not a dash-surrounded marker)."""
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    s = str(v).strip()
    if is_dash_value(v):
        return False
    # String like '-' or '--'
    if s in ("-", "--", ""):
        return False
    # Try to parse as number (handles comma decimals)
    try:
        float(s.replace(",", ".").replace("*", ""))
        return True
    except ValueError:
        return False


def parse_weight(v):
    """Parse a weight value, handling commas and asterisks."""
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".").replace("*", "")
    return float(s)


def year_week_to_date(yw):
    """Convert '2025-27' to a Monday date."""
    parts = yw.split("-")
    year = int(parts[0])
    week = int(parts[1])
    # ISO week date: Monday of the given week
    jan4 = datetime(year, 1, 4)
    start_of_week1 = jan4 - timedelta(days=jan4.weekday())
    return start_of_week1 + timedelta(weeks=week - 1)


def find_col(header_vals, name):
    """Find column index by header name."""
    for i, v in enumerate(header_vals):
        if v and isinstance(v, str) and v.strip() == name:
            return i
    return None


def parse_guide_exercises(ws):
    """Parse the Guide/Pull sheet to extract normal and deload exercise prescriptions."""
    normal_exercises = []
    deload_exercises = []

    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        vals = [cell.value for cell in row]
        all_rows.append(vals)

    # Find header rows (contain "Exercise" as a cell value)
    header_rows = []
    for i, vals in enumerate(all_rows):
        for v in vals:
            if v == "Exercise":
                header_rows.append(i)
                break

    if len(header_rows) < 2:
        print(f"  WARNING: Found {len(header_rows)} header rows, expected 2")
        return [], []

    # Build column map from first header
    def build_col_map(header_vals):
        cols = {}
        for i, v in enumerate(header_vals):
            if v and isinstance(v, str):
                cols[v.strip()] = i
        return cols

    # Parse exercise rows using header-based column mapping
    def parse_section(start_idx, end_idx, header_vals):
        col_map = build_col_map(header_vals)
        ex_col = col_map.get("Exercise", 3)
        ws_col = col_map.get("Working Sets", 8)
        min_col = col_map.get("MIN Reps", 10)
        max_col = col_map.get("MAX Reps", 11)
        early_col = col_map.get("Early Set RPE", 12)
        last_col = col_map.get("Last Set RPE", 13)
        rest_col = col_map.get("REST", 14)
        it_col = col_map.get("Last Set Insensity Technique", 15)
        wu_min_col = col_map.get("MIN Warm-Up Sets Max", 17)
        wu_max_col = col_map.get("MAX Warm-Up Sets", 18)
        sub1_col = col_map.get("Substitution 1", 19)
        sub2_col = col_map.get("Substitution 2", 22)

        exercises = []
        for i in range(start_idx, end_idx):
            vals = all_rows[i]
            name = vals[ex_col] if ex_col < len(vals) else None
            if not name or not isinstance(name, str) or name.strip() in ("Exercise", "Tracking Log", ""):
                continue
            name = name.strip()

            def safe_get(col, default=None):
                if col is not None and col < len(vals):
                    return vals[col]
                return default

            working_sets = safe_get(ws_col)
            min_reps = safe_get(min_col)
            max_reps = safe_get(max_col)
            early_rpe_str = str(safe_get(early_col, "N/A"))
            last_rpe_str = str(safe_get(last_col, "N/A"))
            rest = str(safe_get(rest_col, "2-3 mins"))
            intensity = safe_get(it_col)
            warmup_min = safe_get(wu_min_col, 1)
            warmup_max = safe_get(wu_max_col, 2)

            early_min, early_max = parse_rpe(early_rpe_str)
            last_min, last_max = parse_rpe(last_rpe_str)

            sub1 = None
            sub2 = None
            s1 = safe_get(sub1_col)
            s2 = safe_get(sub2_col)
            if s1 and isinstance(s1, str):
                sub1 = re.sub(r"\s*\(Video\)\s*$", "", s1).strip()
            if s2 and isinstance(s2, str):
                sub2 = re.sub(r"\s*\(Video\)\s*$", "", s2).strip()

            exercises.append({
                "name": name,
                "working_sets": int(float(working_sets)) if working_sets else 2,
                "min_reps": int(min_reps) if min_reps else 8,
                "max_reps": int(max_reps) if max_reps else 10,
                "early_rpe_min": Decimal(str(early_min)) if early_min else Decimal("7"),
                "early_rpe_max": Decimal(str(early_max)) if early_max else Decimal("8"),
                "last_rpe_min": Decimal(str(last_min)) if last_min else Decimal("8"),
                "last_rpe_max": Decimal(str(last_max)) if last_max else Decimal("9"),
                "rest": rest,
                "intensity": str(intensity) if intensity and str(intensity) != "N/A" else None,
                "warmup_min": int(float(warmup_min)) if warmup_min else 1,
                "warmup_max": int(float(warmup_max)) if warmup_max else 2,
                "sub1": sub1,
                "sub2": sub2,
            })
        return exercises

    # Section 1 (normal): between header_rows[0] and header_rows[1]
    normal_exercises = parse_section(header_rows[0] + 1, header_rows[1], all_rows[header_rows[0]])
    # Section 2 (deload): after header_rows[1]
    deload_exercises = parse_section(header_rows[1] + 1, len(all_rows), all_rows[header_rows[1]])

    return normal_exercises, deload_exercises


def parse_exercise_history(ws, exercise_name):
    """Parse an exercise sheet to extract historical set data."""
    history = []  # list of {year_week, sets: [{weight, reps}]}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        vals = [cell.value for cell in row]
        yw = vals[0]
        if not yw or not isinstance(yw, str) or not re.match(r"^\d{4}-\d+$", str(yw)):
            continue

        sets = []
        for s in range(4):  # up to 4 sets
            w_col = 1 + s * 2
            r_col = 2 + s * 2
            weight_val = vals[w_col] if w_col < len(vals) else None
            reps_val = vals[r_col] if r_col < len(vals) else None

            if weight_val is None and reps_val is None:
                continue

            if not is_valid_weight(weight_val):
                continue  # dash-surrounded or invalid → skip this set

            if reps_val is None or not isinstance(reps_val, (int, float)):
                continue

            weight = parse_weight(weight_val)
            reps = int(reps_val)
            if reps <= 0:
                continue

            sets.append({"weight": weight, "reps": reps})

        if sets:
            history.append({"year_week": yw, "sets": sets})

    return history


def main():
    # Collect all data
    all_exercise_names = set()
    all_substitutions = []  # (exercise_name, sub_name, priority)
    exercise_youtube = {}  # exercise_name -> youtube_url
    exercise_notes = {}   # exercise_name -> notes text
    templates = []
    historical_data = []  # (template_name, exercise_name, history)

    for file_cfg in FILES:
        # Load with data_only=False first to get hyperlinks
        wb_links = openpyxl.load_workbook(file_cfg["path"])
        guide_sheet = file_cfg["guide_sheet"]
        template_name = file_cfg["template_name"]

        # Extract exercise notes from cell comments
        if guide_sheet in wb_links.sheetnames:
            ws_notes = wb_links[guide_sheet]
            for row in ws_notes.iter_rows(min_row=1, max_row=ws_notes.max_row):
                for cell in row:
                    if cell.comment:
                        d_cell = ws_notes.cell(row=cell.row, column=4)
                        if d_cell.value and isinstance(d_cell.value, str):
                            name = d_cell.value.strip()
                            text = cell.comment.text
                            text = re.sub(
                                r'\[Threaded comment\].*?Comment:\s*',
                                '', text, flags=re.DOTALL
                            ).strip()
                            # Clean up "Reply:\n\t..." to "Note: ..."
                            text = re.sub(r'\nReply:\s*\n?\t?', '\nNote: ', text)
                            if name and name not in exercise_notes:
                                exercise_notes[name] = text

        # Extract YouTube links from hyperlinks
        if guide_sheet in wb_links.sheetnames:
            ws_links = wb_links[guide_sheet]
            for row in ws_links.iter_rows(min_row=1, max_row=ws_links.max_row):
                for cell in row:
                    if cell.hyperlink and cell.hyperlink.target and "youtube.com" in cell.hyperlink.target:
                        name = str(cell.value).strip() if cell.value else ""
                        name = re.sub(r"\s*\(Video\)\s*$", "", name).strip()
                        name = re.sub(r"\(Video\)$", "", name).strip()
                        if name and name not in exercise_youtube:
                            exercise_youtube[name] = cell.hyperlink.target

        # Load with data_only=True for values
        wb = openpyxl.load_workbook(file_cfg["path"], data_only=True)

        if guide_sheet not in wb.sheetnames:
            print(f"WARNING: {guide_sheet} not found in {file_cfg['path']}")
            continue

        ws = wb[guide_sheet]
        normal_ex, deload_ex = parse_guide_exercises(ws)

        templates.append({
            "name": template_name,
            "normal": normal_ex,
            "deload": deload_ex,
        })

        # Collect exercise names and substitutions
        for ex in normal_ex + deload_ex:
            all_exercise_names.add(ex["name"])
            if ex.get("sub1"):
                all_exercise_names.add(ex["sub1"])
                all_substitutions.append((ex["name"], ex["sub1"], 1))
            if ex.get("sub2"):
                all_exercise_names.add(ex["sub2"])
                all_substitutions.append((ex["name"], ex["sub2"], 2))

        # Parse exercise history sheets
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            # Map sheet name to exercise name
            ex_name = SHEET_NAME_MAP.get(sheet_name, sheet_name)
            all_exercise_names.add(ex_name)

            ws_ex = wb[sheet_name]
            history = parse_exercise_history(ws_ex, ex_name)
            if history:
                historical_data.append((template_name, ex_name, history))

    # De-duplicate substitutions
    seen_subs = set()
    unique_subs = []
    for ex, sub, pri in all_substitutions:
        key = (ex, sub)
        if key not in seen_subs:
            seen_subs.add(key)
            unique_subs.append((ex, sub, pri))

    # ── Generate SQL ────────────────────────────────────────────────────
    sql_lines = []
    sql_lines.append(f"-- Gym Tracker Historical Data Import")
    sql_lines.append(f"-- Generated {datetime.now().isoformat()}")
    sql_lines.append(f"-- User: {USER_EMAIL}")
    sql_lines.append(f"")
    sql_lines.append(f"BEGIN;")
    sql_lines.append(f"")

    # 1. Insert exercises (with youtube_url if available)
    sql_lines.append(f"-- ═══ EXERCISES ═══")
    exercises_with_links = 0
    for name in sorted(all_exercise_names):
        info = EXERCISE_INFO.get(name)
        if not info:
            print(f"WARNING: No muscle_group/equipment for exercise '{name}', defaulting to Other/Other")
            info = ("Other", "Other")
        muscle_group, equipment = info
        yt_url = exercise_youtube.get(name)
        note = exercise_notes.get(name)
        if yt_url:
            exercises_with_links += 1
        sql_lines.append(
            f"INSERT INTO {SCHEMA}.exercises (id, user_id, name, muscle_group, equipment, is_custom, youtube_url, notes, created_at)"
            f" SELECT gen_random_uuid(), NULL, {sql_str(name)}, {sql_str(muscle_group)}, {sql_str(equipment)}, false, {sql_str(yt_url) if yt_url else 'NULL'}, {sql_str(note) if note else 'NULL'}, NOW()"
            f" WHERE NOT EXISTS (SELECT 1 FROM {SCHEMA}.exercises WHERE name = {sql_str(name)} AND user_id IS NULL);"
        )
    sql_lines.append(f"")
    print(f"  Exercises with YouTube links: {exercises_with_links}")
    print(f"  Exercises with notes: {sum(1 for n in all_exercise_names if n in exercise_notes)}")

    # 2. Insert substitutions
    sql_lines.append(f"-- ═══ SUBSTITUTIONS ═══")
    for ex_name, sub_name, priority in unique_subs:
        sql_lines.append(
            f"INSERT INTO {SCHEMA}.exercise_substitutions (id, exercise_id, substitute_exercise_id, priority)"
            f" SELECT gen_random_uuid(),"
            f" (SELECT id FROM {SCHEMA}.exercises WHERE name = {sql_str(ex_name)} AND user_id IS NULL LIMIT 1),"
            f" (SELECT id FROM {SCHEMA}.exercises WHERE name = {sql_str(sub_name)} AND user_id IS NULL LIMIT 1),"
            f" {priority}"
            f" WHERE NOT EXISTS ("
            f"   SELECT 1 FROM {SCHEMA}.exercise_substitutions"
            f"   WHERE exercise_id = (SELECT id FROM {SCHEMA}.exercises WHERE name = {sql_str(ex_name)} AND user_id IS NULL LIMIT 1)"
            f"   AND substitute_exercise_id = (SELECT id FROM {SCHEMA}.exercises WHERE name = {sql_str(sub_name)} AND user_id IS NULL LIMIT 1)"
            f" );"
        )
    sql_lines.append(f"")

    # 3. Insert templates
    sql_lines.append(f"-- ═══ TEMPLATES ═══")
    for tmpl in templates:
        tmpl_name = tmpl["name"]
        sql_lines.append(
            f"INSERT INTO {SCHEMA}.workout_templates (id, user_id, name, created_at)"
            f" SELECT gen_random_uuid(),"
            f" (SELECT id FROM {SCHEMA}.users WHERE email = {sql_str(USER_EMAIL)} LIMIT 1),"
            f" {sql_str(tmpl_name)}, NOW()"
            f" WHERE NOT EXISTS ("
            f"   SELECT 1 FROM {SCHEMA}.workout_templates WHERE name = {sql_str(tmpl_name)}"
            f"   AND user_id = (SELECT id FROM {SCHEMA}.users WHERE email = {sql_str(USER_EMAIL)} LIMIT 1)"
            f" );"
        )
    sql_lines.append(f"")

    # 4. Insert template_exercises
    sql_lines.append(f"-- ═══ TEMPLATE EXERCISES ═══")
    for tmpl in templates:
        tmpl_name = tmpl["name"]
        for week_type, exercises in [("normal", tmpl["normal"]), ("deload", tmpl["deload"])]:
            for order, ex in enumerate(exercises, 1):
                # Average warmup sets
                warmup = (ex["warmup_min"] + ex["warmup_max"]) // 2
                if warmup < 1:
                    warmup = 1

                sql_lines.append(
                    f"INSERT INTO {SCHEMA}.template_exercises"
                    f" (id, template_id, exercise_id, week_type, \"order\", working_sets,"
                    f" min_reps, max_reps, early_set_rpe_min, early_set_rpe_max,"
                    f" last_set_rpe_min, last_set_rpe_max, rest_period, intensity_technique, warmup_sets)"
                    f" SELECT gen_random_uuid(),"
                    f" (SELECT id FROM {SCHEMA}.workout_templates WHERE name = {sql_str(tmpl_name)}"
                    f"  AND user_id = (SELECT id FROM {SCHEMA}.users WHERE email = {sql_str(USER_EMAIL)} LIMIT 1) LIMIT 1),"
                    f" (SELECT id FROM {SCHEMA}.exercises WHERE name = {sql_str(ex['name'])} AND user_id IS NULL LIMIT 1),"
                    f" {sql_str(week_type)}, {order}, {ex['working_sets']},"
                    f" {ex['min_reps']}, {ex['max_reps']},"
                    f" {sql_num(ex['early_rpe_min'])}, {sql_num(ex['early_rpe_max'])},"
                    f" {sql_num(ex['last_rpe_min'])}, {sql_num(ex['last_rpe_max'])},"
                    f" {sql_str(ex['rest'])}, {sql_str(ex['intensity']) if ex['intensity'] else 'NULL'}, {warmup}"
                    f" WHERE NOT EXISTS ("
                    f"   SELECT 1 FROM {SCHEMA}.template_exercises te"
                    f"   JOIN {SCHEMA}.workout_templates wt ON te.template_id = wt.id"
                    f"   WHERE wt.name = {sql_str(tmpl_name)}"
                    f"   AND te.exercise_id = (SELECT id FROM {SCHEMA}.exercises WHERE name = {sql_str(ex['name'])} AND user_id IS NULL LIMIT 1)"
                    f"   AND te.week_type = {sql_str(week_type)}"
                    f"   AND wt.user_id = (SELECT id FROM {SCHEMA}.users WHERE email = {sql_str(USER_EMAIL)} LIMIT 1)"
                    f" );"
                )
    sql_lines.append(f"")

    # 5. Insert historical sessions + sets + progress
    sql_lines.append(f"-- ═══ HISTORICAL WORKOUT DATA ═══")

    # Group history by (template_name, year_week) → create one session per group
    from collections import defaultdict
    sessions = defaultdict(list)  # (template_name, year_week) → [(exercise_name, sets)]
    for template_name, ex_name, history in historical_data:
        for entry in history:
            yw = entry["year_week"]
            sessions[(template_name, yw)].append((ex_name, entry["sets"]))

    for (tmpl_name, yw), exercises_data_list in sorted(sessions.items()):
        session_date = year_week_to_date(yw)
        session_id = str(uuid.uuid4())

        # Create session
        sql_lines.append(f"")
        sql_lines.append(f"-- Session: {tmpl_name} week {yw}")
        sql_lines.append(
            f"INSERT INTO {SCHEMA}.workout_sessions"
            f" (id, user_id, template_id, year_week, week_type, started_at, finished_at, synced)"
            f" SELECT {sql_str(session_id)},"
            f" (SELECT id FROM {SCHEMA}.users WHERE email = {sql_str(USER_EMAIL)} LIMIT 1),"
            f" (SELECT id FROM {SCHEMA}.workout_templates WHERE name = {sql_str(tmpl_name)}"
            f"  AND user_id = (SELECT id FROM {SCHEMA}.users WHERE email = {sql_str(USER_EMAIL)} LIMIT 1) LIMIT 1),"
            f" {sql_str(yw)}, 'normal',"
            f" {sql_str(session_date.strftime('%Y-%m-%d %H:%M:%S'))},"
            f" {sql_str((session_date + timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S'))},"
            f" true"
            f" WHERE NOT EXISTS ("
            f"   SELECT 1 FROM {SCHEMA}.workout_sessions"
            f"   WHERE year_week = {sql_str(yw)}"
            f"   AND template_id = (SELECT id FROM {SCHEMA}.workout_templates WHERE name = {sql_str(tmpl_name)}"
            f"     AND user_id = (SELECT id FROM {SCHEMA}.users WHERE email = {sql_str(USER_EMAIL)} LIMIT 1) LIMIT 1)"
            f"   AND user_id = (SELECT id FROM {SCHEMA}.users WHERE email = {sql_str(USER_EMAIL)} LIMIT 1)"
            f" );"
        )

        # Create sets
        for ex_name, sets_data in exercises_data_list:
            max_weight = max(abs(s["weight"]) for s in sets_data)
            for set_num, s in enumerate(sets_data, 1):
                set_id = str(uuid.uuid4())
                sql_lines.append(
                    f"INSERT INTO {SCHEMA}.workout_sets"
                    f" (id, session_id, exercise_id, set_type, set_number, reps, weight, created_at)"
                    f" SELECT {sql_str(set_id)},"
                    f" {sql_str(session_id)},"
                    f" (SELECT id FROM {SCHEMA}.exercises WHERE name = {sql_str(ex_name)} AND user_id IS NULL LIMIT 1),"
                    f" 'working', {set_num}, {s['reps']}, {s['weight']},"
                    f" {sql_str(session_date.strftime('%Y-%m-%d %H:%M:%S'))}"
                    f" WHERE EXISTS (SELECT 1 FROM {SCHEMA}.workout_sessions WHERE id = {sql_str(session_id)});"
                )

            # Exercise progress (max weight for this exercise this week)
            sql_lines.append(
                f"INSERT INTO {SCHEMA}.exercise_progress"
                f" (id, user_id, exercise_id, year_week, max_weight, created_at)"
                f" SELECT gen_random_uuid(),"
                f" (SELECT id FROM {SCHEMA}.users WHERE email = {sql_str(USER_EMAIL)} LIMIT 1),"
                f" (SELECT id FROM {SCHEMA}.exercises WHERE name = {sql_str(ex_name)} AND user_id IS NULL LIMIT 1),"
                f" {sql_str(yw)}, {max_weight},"
                f" {sql_str(session_date.strftime('%Y-%m-%d %H:%M:%S'))}"
                f" WHERE NOT EXISTS ("
                f"   SELECT 1 FROM {SCHEMA}.exercise_progress"
                f"   WHERE user_id = (SELECT id FROM {SCHEMA}.users WHERE email = {sql_str(USER_EMAIL)} LIMIT 1)"
                f"   AND exercise_id = (SELECT id FROM {SCHEMA}.exercises WHERE name = {sql_str(ex_name)} AND user_id IS NULL LIMIT 1)"
                f"   AND year_week = {sql_str(yw)}"
                f" )"
                f" ON CONFLICT ON CONSTRAINT uq_progress DO NOTHING;"
            )

    sql_lines.append(f"")
    sql_lines.append(f"COMMIT;")

    # Write SQL file
    output_path = r"C:\Users\ZureraA\Personal\gym-app\import_historical_data.sql"
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(sql_lines))

    print(f"SQL file written to: {output_path}")
    print(f"  Exercises: {len(all_exercise_names)}")
    print(f"  Substitutions: {len(unique_subs)}")
    print(f"  Templates: {len(templates)}")
    print(f"  Sessions: {len(sessions)}")
    total_sets = sum(
        sum(len(sets) for _, sets in exercises)
        for exercises in sessions.values()
    )
    print(f"  Total sets: {total_sets}")


if __name__ == "__main__":
    main()
